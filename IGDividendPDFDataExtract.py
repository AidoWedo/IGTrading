"""
This Python script processes a directory of PDF files to extract specific financial data,
namely dividend details and writes this information into an Excel spreadsheet. The script
is designed to handle PDFs containing information on stock dividends, where each PDF
follows a consistent format with the relevant data located on page 3 (In my experience, this is always the case; page 3 holds the dividend information. Obviously if you have 100K shares this might be page 3 to x).

The main functions of the script are:
1. `extract_dividend_details`: This function opens each PDF file, extracts the text from
   page 3, and uses regular expressions to find and extract the date, company name,
   dividend details, and the amount credited.
2. `update_excel`: This function updates an Excel spreadsheet with the extracted data. If
   The spreadsheet does not exist; it is created as a new one with appropriate headers.
3. `process_pdfs`: This function iterates through all PDF files in the specified directory,
   calls the `extract_dividend_details` function for each PDF and collects the extracted
   data. It then calls `update_excel` to write the collected data to the Excel spreadsheet.

The script ensures that the extracted data is written as numerical values where appropriate,
specifically for the dividend amounts, to facilitate any further data analysis or processing
in the Excel file.

Usage:
1. Set the `pdf_directory` variable to the directory's path containing the PDF files.
2. Set the `excel_file_path` variable to the path where the Excel file should be saved.
3. Run the script to process the PDFs and update the Excel file with the extracted dividend details.
"""

import os
import re
import PyPDF2
from openpyxl import load_workbook, Workbook


# Function to extract dividend details from page 3 of a PDF
def extract_dividend_details(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)

        # Check if the PDF has at least 3 pages
        if len(reader.pages) < 3:
            print(f"PDF {pdf_path} does not have at least 3 pages.")
            return None

        # Extract text from page 3
        page = reader.pages[2]
        text = page.extract_text()

        # Debug print to check extracted text
        print(f"Text from page 3 of {pdf_path}:\n{text}\n")

        # Extract details using refined regex
        date_match = re.search(r'(\d{2}[A-Za-z]{3}\d{2})', text)
        name_match = re.search(r'(\b[A-Za-z\s&-]+(Ltd|PLC|Inc|Corp|Co|Group)\b)', text)
        dividend_match = re.search(r'(\d+@\d+\.\d+)', text)
        amount_match = re.search(r'Dividend\s(\d+\.\d+)', text)

        if date_match and name_match and dividend_match and amount_match:
            date = date_match.group(0)
            name = name_match.group(0)
            dividend_details = dividend_match.group(1)
            amount = float(amount_match.group(1))  # Convert amount to float

            # Debug print to check extracted details
            print(
                f"Extracted details - Date: {date}, Name: {name}, Dividend Details: {dividend_details}, Amount: {amount}\n")

            return date, name, dividend_details, amount
    return None


# Function to update Excel sheet
def update_excel(file_path, data):
    if not os.path.exists(file_path):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dividends"
        sheet.append(["Date", "Name", "Dividend Details", "Amount"])
    else:
        workbook = load_workbook(file_path)
        sheet = workbook["Dividends"]

    for entry in data:
        sheet.append(entry)

    workbook.save(file_path)


# Main function to process PDF files in a directory
def process_pdfs(directory, excel_path):
    dividend_data = []

    for filename in os.listdir(directory):
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(directory, filename)
            details = extract_dividend_details(pdf_path)
            if details:
                dividend_data.append(details)

    # Debug print to check collected dividend data uncomment the line below if your missing data to see what is being collected
    # print(f"Collected dividend data: {dividend_data}")

    update_excel(excel_path, dividend_data)


# Directory containing PDFs and the path to the Excel file
pdf_directory = "/path/to/pdf_directory"  # Update to your path i.e. "E:/Documents/Personal/xxxx/"
excel_file_path = "/path/to/output_excel_file.xlsx"  # Update to your path i.e. "E:/Documents/Personal/xxxx/"

# Run the processing
process_pdfs(pdf_directory, excel_file_path)
